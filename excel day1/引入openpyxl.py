import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\prominent\Desktop\用户行为日志0501-0512.xlsx')
# Getting sheets from the workbook
print(wb.sheetnames)
# 也可以按照下面这种方法列出表单
# for sheet in wb:
#     print(sheet.title)

# mysheet = wb.create_sheet('mysheet')  # 也可以用这种方法来添加表单

# 用以下两种方法来指定对某个表单进行操作
# sheet3 = wb.get_sheet_by_name('pro用户行为日志20210501-20210512')
# sheet4 = wb['miot用户行为日志关键字对照表']

# 可以用以下方法对活动表单进行选取
ws = wb.active
# Getting cells from the sheets
print(ws)
print(ws['A3'])
print(ws['A3'].value)

c = ws['B3']
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))
print('Cell{} is {}'.format(c.coordinate, c.value))