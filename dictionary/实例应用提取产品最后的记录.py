import openpyxl
wb1 = openpyxl.Workbook()
wb1.save('销售表.xlsx')
wb = openpyxl.load_workbook('销售表.xlsx')
ws = wb.worksheets[0]
r = 0;d = dict()
while r<ws.max_row-1:
    key = ws.cell_value(r,0)
    value = ws.cell_value(r,1)
    d[key] = value
    r+=1
print(d)