import openpyxl
wb = openpyxl.load_workbook('085400排名.xlsx')
# wb.create_sheet('new build')  # 指定工作表名进行新建，多次运行会进行迭代，new build1,2,3
# wb.create_sheet()  # 会自动生成名字sheet，如果多次运行会进行迭代，sheet1，2，3
wb.create_sheet('指定位置', 2)  # 这样就会指定到第三个位置去创建工作表
wb.save('085400排名.xlsx')
